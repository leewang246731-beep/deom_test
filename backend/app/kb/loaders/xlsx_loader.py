"""XLSX Loader — using openpyxl, preserving headers as context."""
from app.kb.loaders.base_loader import BaseDocumentLoader, LoadedDocument


class XlsxLoader(BaseDocumentLoader):
    @staticmethod
    def supported_extensions() -> list[str]:
        return [".xlsx", ".xls"]

    def load(self, file_path: str) -> LoadedDocument:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_rows = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows.append(f"[Sheet: {sheet_name}]")

            # Extract headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))

            for row in ws.iter_rows(min_row=2, values_only=True):
                parts = []
                for i, val in enumerate(row):
                    if val is not None:
                        header = headers[i] if i < len(headers) else f"Col{i}"
                        parts.append(f"{header}: {val}")
                if parts:
                    all_rows.append("; ".join(parts))

        full_text = "\n".join(all_rows)
        return LoadedDocument(
            text=full_text,
            metadata={
                "source": file_path,
                "format": "xlsx",
                "sheets": wb.sheetnames,
            },
        )
